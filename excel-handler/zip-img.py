from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import os
import sys
from datetime import datetime
import wx
import platform
import warnings
from PIL import ImageFile
import psutil

# 忽略 Pillow 的警告
warnings.filterwarnings('ignore', category=UserWarning)
# 允许截断的图片文件
ImageFile.LOAD_TRUNCATED_IMAGES = True

class CompressionMode:
    """压缩模式"""
    QUALITY = "质量优先"  # 保持较高质量，文件可能较大
    SIZE = "体积优先"     # 优先确保文件小于目标大小
    BALANCED = "平衡模式"  # 在质量和大小之间取平衡

class ExcelImageCompressor:
    """
    Excel图片压缩工具
    用于压缩Excel文件中的图片并生成新的Excel文件
    """
    def __init__(self, log_callback=None):
        # 检测操作系统
        self.is_windows = platform.system().lower() == 'windows'
        # 默认压缩设置
        self.compression_settings = {
            CompressionMode.QUALITY: {"min_quality": 60, "max_size_kb": 500},
            CompressionMode.SIZE: {"min_quality": 5, "max_size_kb": 200},
            CompressionMode.BALANCED: {"min_quality": 30, "max_size_kb": 300}
        }
        self.log_callback = log_callback
        
    def log(self, message):
        """输出日志"""
        print(message)  # 保持控制台输出
        if self.log_callback:
            wx.CallAfter(self.log_callback, message + '\n')
            wx.WakeUpIdle()
        
    def normalize_path(self, path):
        """标准化文件路径，处理不同操作系统的路径差异"""
        return os.path.normpath(path)

    def calculate_new_size(self, width, height, max_size_kb):
        """计算保持宽高比的新尺寸"""
        # 估算目标像素数（基于经验值）
        target_pixels = max_size_kb * 1024 * 0.5  # 0.5是经验系数
        current_pixels = width * height
        
        if current_pixels <= target_pixels:
            return width, height
            
        scale = (target_pixels / current_pixels) ** 0.5
        return int(width * scale), int(height * scale)

    def compress_image(self, image_data, max_size_kb=200, compression_mode=CompressionMode.BALANCED):
        """
        压缩图片数据
        """
        try:
            settings = self.compression_settings[compression_mode]
            min_quality = settings["min_quality"]
            max_size_kb = settings["max_size_kb"]

            # 打开图片
            if isinstance(image_data, io.BytesIO):
                image_data = image_data.getvalue()
            
            # 使用 PIL 打开图片并完全重新创建一个新图片，移除所有元数据
            with PILImage.open(io.BytesIO(image_data)) as img:
                # 创建一个全新的RGB图片
                if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
                    # 处理透明图片
                    new_img = PILImage.new('RGB', img.size, (255, 255, 255))
                    if img.mode == 'RGBA':
                        new_img.paste(img, mask=img.split()[3])
                    else:
                        new_img.paste(img)
                else:
                    # 直接转换为RGB
                    new_img = PILImage.new('RGB', img.size, (255, 255, 255))
                    new_img.paste(img)

                # 调整大小（如果需要）
                if new_img.size[0] * new_img.size[1] > 4000 * 3000:
                    ratio = min(4000/new_img.size[0], 3000/new_img.size[1])
                    new_size = (int(new_img.size[0] * ratio), int(new_img.size[1] * ratio))
                    new_img = new_img.resize(new_size, PILImage.LANCZOS)

                # 设置压缩质量
                quality = 95 if compression_mode == CompressionMode.QUALITY else \
                         85 if compression_mode == CompressionMode.BALANCED else 70

                # 保存图片
                output = io.BytesIO()
                new_img.save(output, 
                            format='JPEG',
                            quality=quality,
                            optimize=True,
                            progressive=True)

                # 如果文件仍然太大，继续降低质量
                while len(output.getvalue()) > max_size_kb * 1024 and quality > min_quality:
                    quality -= 5
                    output = io.BytesIO()
                    new_img.save(output,
                               format='JPEG',
                               quality=quality,
                               optimize=True,
                               progressive=True)

                return output.getvalue()

        except Exception as e:
            raise Exception(f"压缩图片时出错: {str(e)}")

    def process_excel(self, file_path, compression_mode=CompressionMode.BALANCED):
        try:
            # 添加诊断信息
            file_size = os.path.getsize(file_path)
            free_space = psutil.disk_usage(os.path.dirname(file_path)).free
            self.log(f"原始文件大小: {file_size/1024/1024:.2f}MB")
            self.log(f"磁盘剩余空间: {free_space/1024/1024:.2f}MB")
            self.log(f"Python版本: {sys.version}")
            self.log(f"操作系统: {platform.platform()}")
            
            file_path = self.normalize_path(file_path)
            self.log(f"正在处理文件：{file_path}")
            
            # 先尝试读取文件，��保文件可以正常打开
            try:
                wb = load_workbook(file_path, keep_vba=True, data_only=False, keep_links=True)
            except Exception as e:
                raise Exception(f"无法打开Excel文件：{str(e)}")

            processed_images = 0
            total_images = 0
            
            # 保存原始工作簿的一些重要属性
            excel_properties = {
                'encoding': 'utf-8',
                'has_properties': hasattr(wb, 'properties'),
                'has_vba': wb.vba_archive if hasattr(wb, 'vba_archive') else None,
            }
            
            # 处理每个工作表
            for sheet in wb.worksheets:
                images = sheet._images.copy()
                total_images += len(images)
                
                if not images:
                    continue
                
                image_info = []
                for img in images:
                    try:
                        # 获取详细的锚点信息
                        anchor_info = {}
                        if hasattr(img.anchor, '_from'):
                            anchor_from = img.anchor._from
                            anchor_info['from'] = {
                                'col': anchor_from.col,
                                'row': anchor_from.row,
                                'colOff': anchor_from.colOff,
                                'rowOff': anchor_from.rowOff
                            }
                        if hasattr(img.anchor, 'to'):
                            anchor_to = img.anchor.to
                            anchor_info['to'] = {
                                'col': anchor_to.col,
                                'row': anchor_to.row,
                                'colOff': anchor_to.colOff,
                                'rowOff': anchor_to.rowOff
                            }
                        
                        image_info.append({
                            'image': img,
                            'anchor_type': img.anchor.__class__.__name__,
                            'anchor_info': anchor_info,
                            'width': img.width,
                            'height': img.height,
                            'ext': img.anchor.ext if hasattr(img.anchor, 'ext') else None
                        })
                    except Exception as e:
                        self.log(f"警告：获取���片信息时出错，将保持原图：{str(e)}")
                        image_info.append({'image': img, 'use_original': True})
                
                # 清除并重新添加图片
                sheet._images.clear()
                
                for i, info in enumerate(image_info, 1):
                    try:
                        if info.get('use_original'):
                            sheet.add_image(info['image'])
                            continue

                        img = info['image']
                        if hasattr(img, '_data'):
                            image_data = img._data()
                        else:
                            image_data = img.ref
                        
                        compressed_data = self.compress_image(
                            image_data, 
                            compression_mode=compression_mode
                        )
                        
                        new_img = Image(io.BytesIO(compressed_data))
                        
                        # 设置锚点
                        if info['anchor_type'] == 'OneCellAnchor':
                            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
                            marker = AnchorMarker(**info['anchor_info']['from'])
                            new_img.anchor = OneCellAnchor(_from=marker, ext=info['ext'])
                        elif info['anchor_type'] == 'TwoCellAnchor':
                            from openpyxl.drawing.spreadsheet_drawing import TwoCellAnchor, AnchorMarker
                            marker1 = AnchorMarker(**info['anchor_info']['from'])
                            marker2 = AnchorMarker(**info['anchor_info']['to'])
                            new_img.anchor = TwoCellAnchor(_from=marker1, to=marker2)
                        
                        new_img.width = info['width']
                        new_img.height = info['height']
                        
                        sheet.add_image(new_img)
                        processed_images += 1
                        
                    except Exception as e:
                        self.log(f"处理第 {i} 张图片时出错：{str(e)}")
                        sheet.add_image(info['image'])

            # 生成新文件路径
            file_dir = os.path.dirname(file_path)
            file_name = os.path.basename(file_path)
            name, ext = os.path.splitext(file_name)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            new_file_path = os.path.join(file_dir, f"{name}_compressed_{timestamp}{ext}")
            new_file_path = self.normalize_path(new_file_path)

            # 修改保存文件的逻辑
            try:
                # 先尝试常规保存
                wb.save(new_file_path)
                
                # 验证新文件是否可以正常打开
                test_wb = load_workbook(new_file_path)
                test_wb.close()
                
            except Exception as save_error:
                self.log(f"常规保存失败，尝试使用备选方案：{str(save_error)}")
                
                # 备选保存方案 1：使用 force_zip64
                try:
                    wb.save(new_file_path, force_zip64=True)
                except Exception as e1:
                    self.log(f"备选方案1失败：{str(e1)}")
                    
                    # 备选保存方案 2：重新加载并保存
                    try:
                        wb = load_workbook(file_path, keep_vba=True, data_only=False, keep_links=True)
                        wb.save(new_file_path, force_zip64=True)
                    except Exception as e2:
                        raise Exception(f"所有保存方案都失败：\n1. {str(save_error)}\n2. {str(e1)}\n3. {str(e2)}")

            message = (f"处理完成！\n"
                      f"压缩模式：{compression_mode}\n"
                      f"共处理 {processed_images}/{total_images} 张图片\n"
                      f"新文件保存在：\n{new_file_path}")
            self.log(message)
            wx.MessageBox(message, "成功", wx.OK | wx.ICON_INFORMATION)
            return True
            
        except Exception as e:
            error_message = f"处理文件时出错：\n{str(e)}"
            self.log(error_message)
            wx.MessageBox(error_message, "错误", wx.OK | wx.ICON_ERROR)
            return False

class MainFrame(wx.Frame):
    def __init__(self):
        size = (650, 500) if platform.system().lower() == 'windows' else (600, 500)
        super().__init__(parent=None, title='Excel图片压缩工具', size=size)
        self.init_ui()
        
        # 窗口居中显示
        self.Center()
        
        # 设置窗口样式
        self.SetWindowStyle(
            wx.DEFAULT_FRAME_STYLE | 
            wx.STAY_ON_TOP  # 窗口置顶
        )
        
        # 设置图标（如果需要）
        try:
            icon = wx.Icon("logo_N.png", wx.BITMAP_TYPE_PNG)
            self.SetIcon(icon)
        except:
            pass
            
    def Show(self, show=True):
        """重写Show方法，增加窗口显示效果"""
        super().Show(show)
        if show:
            # 将窗口置为前台
            self.Raise()
            # 设置焦点
            self.SetFocus()
            
            # 闪烁窗口提示用户（仅Windows系统）
            if platform.system().lower() == 'windows':
                self.RequestUserAttention()
            
            # 延迟100ms后取消置顶
            wx.CallLater(100, self.remove_stay_on_top)
    
    def remove_stay_on_top(self):
        """取消窗口置顶"""
        style = self.GetWindowStyle()
        self.SetWindowStyle(style & ~wx.STAY_ON_TOP)
        
    def init_ui(self):
        panel = wx.Panel(self)
        vbox = wx.BoxSizer(wx.VERTICAL)
        
        # 文件选择部分
        label = wx.StaticText(panel, label="选择要处理的Excel文件：")
        vbox.Add(label, 0, wx.ALL | wx.EXPAND, 5)
        
        hbox = wx.BoxSizer(wx.HORIZONTAL)
        self.file_path = wx.TextCtrl(panel)
        browse_btn = wx.Button(panel, label="浏览...")
        hbox.Add(self.file_path, 1, wx.EXPAND | wx.RIGHT, 5)
        hbox.Add(browse_btn, 0)
        vbox.Add(hbox, 0, wx.ALL | wx.EXPAND, 5)
        
        # 压缩模式选择
        mode_label = wx.StaticText(panel, label="选择压缩模式：")
        vbox.Add(mode_label, 0, wx.ALL | wx.EXPAND, 5)
        
        self.mode_choice = wx.Choice(panel, choices=[
            CompressionMode.QUALITY,
            CompressionMode.BALANCED,
            CompressionMode.SIZE
        ])
        self.mode_choice.SetSelection(1)  # 默认选择平衡模式
        vbox.Add(self.mode_choice, 0, wx.ALL | wx.EXPAND, 5)
        
        # 处理按钮
        process_btn = wx.Button(panel, label="开始处理")
        vbox.Add(process_btn, 0, wx.ALL | wx.EXPAND, 5)
        
        # 添加日志文本框
        log_label = wx.StaticText(panel, label="处理日志：")
        vbox.Add(log_label, 0, wx.ALL | wx.EXPAND, 5)
        
        self.log_text = wx.TextCtrl(panel, style=wx.TE_MULTILINE | wx.TE_READONLY | wx.HSCROLL)
        vbox.Add(self.log_text, 1, wx.ALL | wx.EXPAND, 5)
        
        panel.SetSizer(vbox)
        
        # 绑定事件
        browse_btn.Bind(wx.EVT_BUTTON, self.on_browse)
        process_btn.Bind(wx.EVT_BUTTON, self.on_process)
        
    def log(self, message):
        """添加日志到文本框"""
        self.log_text.AppendText(message)
        self.log_text.ShowPosition(self.log_text.GetLastPosition())
        
    def on_browse(self, event):
        with wx.FileDialog(self, "选择Excel文件", wildcard="Excel文件 (*.xlsx;*.xls)|*.xlsx;*.xls",
                          style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST) as fileDialog:
            if fileDialog.ShowModal() == wx.ID_CANCEL:
                return
            self.file_path.SetValue(fileDialog.GetPath())
            
    def on_process(self, event):
        file_path = self.file_path.GetValue()
        if not file_path:
            wx.MessageBox("请先选Excel文件！", "错误", wx.OK | wx.ICON_ERROR)
            return
        
        # 清空日志
        self.log_text.SetValue("")
        
        # 禁用处理按钮，避免重复点击
        event.GetEventObject().Disable()
        
        try:
            compression_mode = self.mode_choice.GetString(self.mode_choice.GetSelection())
            compressor = ExcelImageCompressor(log_callback=self.log)
            
            # 使用线程处理压缩任务
            def process_task():
                try:
                    compressor.process_excel(file_path, compression_mode)
                finally:
                    # 处理完成后重新启用按钮
                    wx.CallAfter(event.GetEventObject().Enable)
            
            import threading
            thread = threading.Thread(target=process_task)
            thread.daemon = True
            thread.start()
            
        except Exception as e:
            wx.MessageBox(f"处理过程中出错：{str(e)}", "错误", wx.OK | wx.ICON_ERROR)
            event.GetEventObject().Enable()

if __name__ == '__main__':
    app = wx.App()
    frame = MainFrame()
    frame.Show()
    app.MainLoop()
